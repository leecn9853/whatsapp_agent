"""成本报表图片渲染的私有 helper：LibreOffice 重算公式、读取看板区块、生成最终图片。

生成最终图片目前实际在用的是两条"原样截图"路径（LibreOffice 转 PDF 再用 pymupdf 渲成图片，
保留 Excel/LibreOffice 原生的边框/底色/合并单元格/图表配色，不是重新排版画的）：
1. render_dashboard_screenshot：截「成本分析看板」sheet 整体（3 个区块——月度趋势/部门预算/
   采购类别——一起截，不按 section 拆；列数少的区块会被居中挪到列数最多的区块下面，而不是
   永远贴在最左边）。
2. render_chart_screenshot：截「成本图表」sheet 整体（3 个原生 Excel 图表一起，不按 section 拆）。

render_table_image / render_chart_image 是更早的实现，把区块数值读出来用 matplotlib 重新排版
画一张表格图/图表图——样式（配色、字号）是这边代码自己控制的，跟 Excel 原本的单元格格式/图表
配色无关。当前 cost_report_tools.py 已经不再调用这两个函数（只保留代码，不接入 render_type），
截图路径信息更真实、跟 Excel 里看到的一致，所以换成了默认路径。

不对 LLM 暴露任何 @tool，只被 cost_report_tools.py 内部调用。
"""
from __future__ import annotations

import io
import os
import subprocess
from copy import copy
from pathlib import Path

import matplotlib

matplotlib.use("Agg")

import pymupdf  # 把 LibreOffice 转出来的 PDF 页面渲成图片
import matplotlib.font_manager as fm
import matplotlib.pyplot as plt
import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from PIL import Image, ImageDraw, ImageFont

SOFFICE_BIN = os.getenv("SOFFICE_BIN", "soffice")
_DASHBOARD_SHEET = "成本分析看板"
_CHART_SHEET = "成本图表"

_PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent.parent
_FONT_PATH = _PROJECT_ROOT / "assets" / "fonts" / "NotoSansSC-Regular.ttf"

_HEADER_COLOR = "#1F4E78"
_BAND_COLOR = "#F2F2F2"
_TOTAL_ROW_COLOR = "#D9E1F2"

_font_ready = False


def _ensure_cjk_font() -> None:
    global _font_ready
    if _font_ready:
        return
    fm.fontManager.addfont(str(_FONT_PATH))
    font_name = fm.FontProperties(fname=str(_FONT_PATH)).get_name()
    plt.rcParams["font.family"] = font_name
    plt.rcParams["axes.unicode_minus"] = False
    _font_ready = True


class CostReportRenderError(Exception):
    """渲染/重算阶段的内部错误，转成给 agent 看的错误文本，不抛异常中断整轮对话。"""


def _soffice_convert(input_path: Path, outdir: Path, profile_dir: Path, fmt: str) -> Path:
    """用 LibreOffice headless 把 input_path 转换成 fmt 格式，返回转换后文件的路径。

    outdir/profile_dir 由调用方提供（调用方负责整体临时目录的生命周期与清理），每次调用要用
    各自独立的目录，避免并发调用时抢同一份 LibreOffice 用户配置锁。recalculate_with_libreoffice
    （fmt="xlsx"，重算公式）和 _convert_to_pdf（fmt="pdf"，配合 print_area 截图某个区块）复用
    这一段 subprocess 调用逻辑，区别只在 --convert-to 传的格式。
    """
    try:
        result = subprocess.run(
            [
                SOFFICE_BIN,
                "--headless",
                "--calc",
                "--convert-to",
                fmt,
                "--outdir",
                str(outdir),
                str(input_path),
                f"-env:UserInstallation=file://{profile_dir}",
            ],
            capture_output=True,
            text=True,
            timeout=60,
        )
    except FileNotFoundError as e:
        raise CostReportRenderError(
            "未找到 soffice 命令，请确认已安装 LibreOffice（brew install libreoffice）。"
        ) from e
    except subprocess.TimeoutExpired as e:
        raise CostReportRenderError("LibreOffice 转换超时。") from e

    if result.returncode != 0:
        raise CostReportRenderError(f"LibreOffice 转换失败：{result.stderr or result.stdout}")

    converted = outdir / f"{input_path.stem}.{fmt}"
    if not converted.exists():
        raise CostReportRenderError("LibreOffice 没有生成转换后的文件。")
    return converted


def recalculate_with_libreoffice(input_path: Path, outdir: Path, profile_dir: Path) -> Path:
    """用 LibreOffice headless 重新计算 input_path 里所有公式，返回重算后文件的路径。

    outdir/profile_dir 由调用方提供（调用方负责整体临时目录的生命周期与清理），每次调用要用
    各自独立的目录，避免并发调用时抢同一份 LibreOffice 用户配置锁。
    """
    return _soffice_convert(input_path, outdir, profile_dir, "xlsx")


def _convert_to_pdf(input_path: Path, outdir: Path, profile_dir: Path) -> Path:
    """把 input_path 转成 PDF——render_section_screenshot 用来把设好 print_area 的 sheet
    导出成「只有一页、内容正好是打印区域」的 PDF，再交给 pymupdf 渲成图片。
    """
    return _soffice_convert(input_path, outdir, profile_dir, "pdf")


def read_section(recalculated_path: Path, section_cfg: dict) -> pd.DataFrame:
    """从重算后的 xlsx 的「成本分析看板」sheet 里读出一个区块（含合计/平均行），转成 DataFrame。"""
    wb = openpyxl.load_workbook(recalculated_path, data_only=True)
    ws = wb[_DASHBOARD_SHEET]
    columns = section_cfg["columns"]
    n_cols = len(columns)
    rows = []
    for row in range(section_cfg["data_first_row"], section_cfg["total_row"] + 1):
        rows.append([ws.cell(row=row, column=c).value for c in range(1, n_cols + 1)])
    df = pd.DataFrame(rows, columns=columns)

    sort_by = section_cfg.get("sort_by")
    if sort_by:
        # 模板「分类整理」sheet 用 UNIQUE() 去重月份/类别，没有排序，保留的是原始数据里
        # 第一次出现的顺序——月份是 YYYY-MM 格式，文本排序即为时间顺序，这里只影响画图
        # 展示顺序，不改模板本身。合计/平均行始终是最后一行，排序时单独摘出来后拼回末尾。
        data_df, total_row = df.iloc[:-1], df.iloc[[-1]]
        df = pd.concat([data_df.sort_values(sort_by).reset_index(drop=True), total_row], ignore_index=True)
    return df


def _format_cell(value, col_name: str, percent_cols: set[str]) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    if col_name in percent_cols:
        return f"{value * 100:.1f}%"
    if float(value).is_integer():
        return f"{int(value):,}"
    return f"{value:,.2f}"


def render_table_image(df: pd.DataFrame, section_cfg: dict, out_path: Path, footer_text: str) -> None:
    """把区块数值用 matplotlib 重新画成一张表格图——不是 Excel 原始单元格的截图，配色/字号/
    底纹都是这里手写的，跟模板里「成本分析看板」sheet 本身的单元格格式无关。原样截图走的是
    render_section_screenshot（同目录）这条完全独立的路径。
    """
    _ensure_cjk_font()
    percent_cols = set(section_cfg.get("percent_cols", []))
    columns = section_cfg["columns"]
    n_rows, n_cols = df.shape

    cell_text = [
        [_format_cell(row[col], col, percent_cols) for col in columns] for _, row in df.iterrows()
    ]

    fig_width = max(6.0, n_cols * 1.7)
    fig_height = max(2.0, (n_rows + 1) * 0.5)
    fig, ax = plt.subplots(figsize=(fig_width, fig_height))
    ax.axis("off")
    ax.set_title(section_cfg["title"], fontsize=14, fontweight="bold", pad=12)

    table = ax.table(cellText=cell_text, colLabels=columns, loc="center", cellLoc="center")
    table.auto_set_font_size(False)
    table.set_fontsize(10)
    table.scale(1, 1.6)
    table.auto_set_column_width(col=list(range(n_cols)))

    for (r, _c), cell in table.get_celld().items():
        if r == 0:
            cell.set_facecolor(_HEADER_COLOR)
            cell.set_text_props(color="white", fontweight="bold")
        elif r == n_rows:
            cell.set_facecolor(_TOTAL_ROW_COLOR)
            cell.set_text_props(fontweight="bold")
        elif r % 2 == 0:
            cell.set_facecolor(_BAND_COLOR)

    fig.text(0.5, 0.01, footer_text, ha="center", va="bottom", fontsize=7, color="#666666")
    fig.tight_layout(rect=(0, 0.04, 1, 1))
    fig.savefig(out_path, dpi=150, bbox_inches="tight")
    plt.close(fig)


def _content_width_units(value) -> float:
    """粗略估算一个单元格值需要的 Excel 列宽（字符宽度单位），中文字符按 1.8 个英文字符宽算。"""
    text = "" if value is None else str(value)
    return sum(1.8 if ord(ch) > 0x2E80 else 1.0 for ch in text)


def _flatten_dashboard_formulas(wb, wb_values) -> None:
    """把「成本分析看板」sheet 里的公式单元格换成 data_only 读出来的静态值。

    截图前通常要把「成本分析看板」依赖的原始明细/分类整理 sheet 删掉（减少 --convert-to pdf
    转出的页数），公式单元格如果还留着公式字符串，删掉被引用的 sheet 后就会变成 #REF!——
    这里提前把公式换成算好的静态值，就不受删 sheet 影响。render_section_screenshot 和
    render_chart_screenshot 都要用到（后者的原生图表直接读这张 sheet 的单元格值画图）。
    """
    ws_values = wb_values[_DASHBOARD_SHEET]
    ws = wb[_DASHBOARD_SHEET]
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.value = ws_values.cell(row=cell.row, column=cell.column).value


def _render_pdf_page_with_footer(pdf_path: Path, page_index: int, out_path: Path, footer_text: str) -> None:
    """用 pymupdf 把 PDF 的第 page_index 页渲成高分辨率 PNG，再用 Pillow 在底部贴一条脚注
    （跟 render_table_image/render_chart_image 的脚注视觉风格保持一致，只是画法从 matplotlib
    换成 Pillow）。render_section_screenshot 和 render_chart_screenshot 共用。
    """
    doc = pymupdf.open(pdf_path)
    pix = doc[page_index].get_pixmap(dpi=200)
    doc.close()
    section_img = Image.open(io.BytesIO(pix.tobytes("png"))).convert("RGB")

    _ensure_cjk_font()
    footer_height = 28
    canvas = Image.new("RGB", (section_img.width, section_img.height + footer_height), "white")
    canvas.paste(section_img, (0, 0))
    draw = ImageDraw.Draw(canvas)
    font = ImageFont.truetype(str(_FONT_PATH), 16)
    text_width = draw.textlength(footer_text, font=font)
    draw.text(
        ((section_img.width - text_width) / 2, section_img.height + 4),
        footer_text,
        font=font,
        fill="#666666",
    )
    canvas.save(out_path)


def _reset_cell_style(cell) -> None:
    cell.value = None
    cell.font = Font()
    cell.fill = PatternFill()
    cell.border = Border()
    cell.alignment = Alignment()
    cell.number_format = "General"


def _move_cell(src, dest) -> None:
    dest.value = src.value
    dest.font = copy(src.font)
    dest.fill = copy(src.fill)
    dest.border = copy(src.border)
    dest.alignment = copy(src.alignment)
    dest.number_format = src.number_format


def render_dashboard_screenshot(
    recalculated_path: Path, sections_cfg: list[dict], out_path: Path, footer_text: str, tmp_dir: Path
) -> None:
    """把「成本分析看板」sheet 整体（3 个区块——标题行到合计行，一个接一个纵向排列）原样
    截图成一张 PNG，不按区块拆开截、也不按哪个区块单独截——保留 Excel/LibreOffice 原生的
    边框、单元格底色、合并单元格，不像 render_table_image 那样重新排版画图。

    做法：
    1. 跟 render_chart_screenshot 一样用 _flatten_dashboard_formulas 把公式换成静态值，删掉
       「成本分析看板」外的所有 sheet，避免 #REF! 和多余的 PDF 页。
    2. 3 个区块列数不一样（月度趋势 8 列、部门预算 4 列、采购类别 7 列），但共享同一张 sheet
       的列宽网格——列数少的区块如果永远贴着 A 列开头，跟旁边列数最多的区块比会明显偏左，看
       起来不居中。这里把列数比最多列数少的区块，其标题行到合计行整块内容（连同单元格格式）
       物理搬到居中的列位置（offset = (最多列数 - 该区块列数) // 2），原列位置清空还原成默认
       格式，避免残留边框/底色。搬移按列号从大到小处理，目标列永远比当前源列大，不会互相
       覆盖还没读到的源列。
    3. 每一列的列宽按「所有区块搬移后落在这一列的表头/数据/合计内容」重新算一遍（不含标题行，
       标题行单独按各自区块的合并宽度补差额，逻辑跟之前单区块截图一致，只是窗口从「第 1 列到
       该区块列数」变成「该区块搬移后的列区间」）。
    4. print_area 设成从第一个区块的标题行到最后一个区块的合计行、宽度覆盖列数最多的区块；
       额外打开 print_options 的 horizontalCentered/verticalCentered——fitToWidth/fitToHeight
       各自独立缩放取较小值，两个方向不一定都刚好撑满一页，居中选项能让撑不满的那个方向的
       留白平均分布在两侧，而不是全部堆在右边/下边。
    5. 转 PDF，取第 0 页（这份临时文件只留了「成本分析看板」一个 sheet），渲成图片贴脚注。

    tmp_dir 由调用方提供并负责清理（跟 recalculate_with_libreoffice 的 outdir/profile_dir
    是同一套约定），这里会在它下面再建两个子目录给 LibreOffice 转 PDF 用。
    """
    wb_values = openpyxl.load_workbook(recalculated_path, data_only=True)
    wb = openpyxl.load_workbook(recalculated_path)

    ws = wb[_DASHBOARD_SHEET]
    _flatten_dashboard_formulas(wb, wb_values)

    for name in list(wb.sheetnames):
        if name != _DASHBOARD_SHEET:
            del wb[name]

    max_n_cols = max(len(cfg["columns"]) for cfg in sections_cfg)

    # 区块的标题行在 data_first_row 往上数两行（一行空行分隔 + 一行区块标题），
    # 见模板「成本分析看板」sheet 的实际排版（第 3/20/31 行是各区块标题）。
    blocks = []
    for cfg in sections_cfg:
        title_row = cfg["data_first_row"] - 2
        n_cols = len(cfg["columns"])
        blocks.append(
            {
                "title_row": title_row,
                "header_row": title_row + 1,
                "total_row": cfg["total_row"],
                "n_cols": n_cols,
                "offset": (max_n_cols - n_cols) // 2,
            }
        )

    for block in blocks:
        offset = block["offset"]
        if offset == 0:
            continue
        for row in range(block["title_row"], block["total_row"] + 1):
            for col in range(block["n_cols"], 0, -1):
                src = ws.cell(row=row, column=col)
                _move_cell(src, ws.cell(row=row, column=col + offset))
                _reset_cell_style(src)

    for col_idx in range(1, max_n_cols + 1):
        widths = [
            _content_width_units(ws.cell(row=r, column=col_idx).value)
            for block in blocks
            for r in range(block["header_row"], block["total_row"] + 1)
        ]
        ws.column_dimensions[get_column_letter(col_idx)].width = max(widths, default=8.0) + 3

    # 列少的区块按数据内容算出来的总列宽可能撑不起标题行那行长文字，合并单元格会在打印区域
    # 边界被截断——这里按标题需要的宽度把差额平摊加回该区块自己占的那几列（*1.2 是给标题的
    # 粗体字号留的余量）。注意标题用的是模板里标题单元格的实际文字（带「一、」编号和括号里的
    # 数据来源说明），不能用 cfg["title"] 那个给 matplotlib 图表用的短标题算宽度。
    for block in blocks:
        offset, n_cols = block["offset"], block["n_cols"]
        first_col, last_col = offset + 1, offset + n_cols
        actual_title = ws.cell(row=block["title_row"], column=first_col).value
        title_units = _content_width_units(actual_title) * 1.2
        total_width = sum(ws.column_dimensions[get_column_letter(c)].width for c in range(first_col, last_col + 1))
        if title_units > total_width:
            extra_per_col = (title_units - total_width) / n_cols
            for c in range(first_col, last_col + 1):
                ws.column_dimensions[get_column_letter(c)].width += extra_per_col
        ws.merge_cells(f"{get_column_letter(first_col)}{block['title_row']}:{get_column_letter(last_col)}{block['title_row']}")

    top_row = blocks[0]["title_row"]
    bottom_row = blocks[-1]["total_row"]
    ws.print_area = f"A{top_row}:{get_column_letter(max_n_cols)}{bottom_row}"
    ws.print_options.gridLines = False
    ws.print_options.headings = False
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    ws.page_margins = PageMargins(left=0.2, right=0.2, top=0.2, bottom=0.2, header=0, footer=0)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    src_path = tmp_dir / "dashboard_screenshot_src.xlsx"
    wb.save(src_path)

    pdf_outdir = tmp_dir / "pdf_out"
    pdf_profile = tmp_dir / "pdf_profile"
    pdf_outdir.mkdir(exist_ok=True)
    pdf_profile.mkdir(exist_ok=True)
    pdf_path = _convert_to_pdf(src_path, pdf_outdir, pdf_profile)

    _render_pdf_page_with_footer(pdf_path, 0, out_path, footer_text)


def render_chart_screenshot(recalculated_path: Path, out_path: Path, footer_text: str, tmp_dir: Path) -> None:
    """把「成本图表」sheet 整体原样截图成 PNG——sheet 里 3 个原生图表（月度趋势/部门预算/
    采购类别）一起截，不按 section 拆开截单独一个图表（如果之后有单独截一个图表的需求再改）。

    图表的数据系列直接引用「成本分析看板」sheet 的单元格（不是原始明细 sheet），所以这份图表
    sheet 离不开「成本分析看板」——不能像 render_section_screenshot 那样把它也删掉，只能把它
    设成隐藏，减少 --convert-to pdf 转出的页数。

    做法：
    1. 跟 render_section_screenshot 一样用 _flatten_dashboard_formulas 把「成本分析看板」的
       公式换成静态值，删掉两张看板 sheet 之外的原始明细/分类整理 sheet（图表不直接依赖它们，
       只是「成本分析看板」的公式在被删前依赖，flatten 过后就安全了）。
    2. 隐藏「成本分析看板」，只留「成本图表」可见；给一个够大的 print_area 覆盖 3 个图表
       （默认列宽/行高下，3 个图表纵向堆叠大约占到 A2:M53，这里留够余量设成 A1:M56）。
    3. 转 PDF 后取最后一页，而不是固定第 0 页——「成本图表」在 workbook 里始终是最后一个
       sheet，不依赖"隐藏的 sheet 一定不会输出页面"这个不同 LibreOffice 版本可能不一致的假设，
       取最后一页更稳。
    """
    wb_values = openpyxl.load_workbook(recalculated_path, data_only=True)
    wb = openpyxl.load_workbook(recalculated_path)

    _flatten_dashboard_formulas(wb, wb_values)

    for name in list(wb.sheetnames):
        if name not in (_DASHBOARD_SHEET, _CHART_SHEET):
            del wb[name]

    wb[_DASHBOARD_SHEET].sheet_state = "hidden"

    ws = wb[_CHART_SHEET]
    ws.print_area = "A1:M56"
    ws.print_options.gridLines = False
    ws.print_options.headings = False
    ws.page_margins = PageMargins(left=0.2, right=0.2, top=0.2, bottom=0.2, header=0, footer=0)
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    src_path = tmp_dir / "chart_screenshot_src.xlsx"
    wb.save(src_path)

    pdf_outdir = tmp_dir / "pdf_out"
    pdf_profile = tmp_dir / "pdf_profile"
    pdf_outdir.mkdir(exist_ok=True)
    pdf_profile.mkdir(exist_ok=True)
    pdf_path = _convert_to_pdf(src_path, pdf_outdir, pdf_profile)

    doc = pymupdf.open(pdf_path)
    last_page = len(doc) - 1
    doc.close()
    _render_pdf_page_with_footer(pdf_path, last_page, out_path, footer_text)


def render_chart_image(df: pd.DataFrame, section_cfg: dict, out_path: Path, footer_text: str) -> None:
    """把区块数值用 matplotlib 重新画成一张折线图/柱状图——同样是重新画的，不是截图，
    见 render_table_image 顶部的说明。
    """
    _ensure_cjk_font()
    chart_cfg = section_cfg["chart"]
    chart_df = df.iloc[:-1]  # 排除「合计/平均」行，不适合和其余数据点混在一张图里

    x_labels = chart_df[chart_cfg["x"]].tolist()
    series_names = chart_cfg["series"]

    fig, ax = plt.subplots(figsize=(max(6.0, len(x_labels) * 0.8), 5.0))

    if chart_cfg["kind"] == "line":
        for name in series_names:
            ax.plot(x_labels, chart_df[name], marker="o", label=name)
    else:
        width = 0.35
        positions = range(len(x_labels))
        for i, name in enumerate(series_names):
            offset = (i - (len(series_names) - 1) / 2) * width
            ax.bar([p + offset for p in positions], chart_df[name], width=width, label=name)
        ax.set_xticks(list(positions))
        ax.set_xticklabels(x_labels)

    ax.set_title(section_cfg["title"], fontsize=14, fontweight="bold")
    ax.set_ylabel("万元")
    ax.legend()
    ax.grid(alpha=0.3)
    plt.setp(ax.get_xticklabels(), rotation=45, ha="right")

    fig.text(0.5, 0.01, footer_text, ha="center", va="bottom", fontsize=7, color="#666666")
    fig.tight_layout(rect=(0, 0.04, 1, 1))
    fig.savefig(out_path, dpi=150, bbox_inches="tight")
    plt.close(fig)
