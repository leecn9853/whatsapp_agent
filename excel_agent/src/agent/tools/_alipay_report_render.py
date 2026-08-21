"""支付宝匹配数据表的私有 helper：金额区间分桶、数据类别/明细分类动态分组聚合、
openpyxl 建表与样式，以及把生成的单一 sheet 原样截图成 PNG（render_screenshot，复用
_report_screenshot.py 里跟成本报表共用的 LibreOffice/PDF/脚注原语）。

不含任何网络请求，不暴露 `@tool`，只被 alipay_report_tools.py 内部调用。**类别/明细分类
的取值集合和显示顺序 100% 由传入的 records 决定，本文件不允许出现任何硬编码的分类列表**——
第三方接口以后可能返回开发者预先不知道的新类别，也不需要跟着改这里的代码。
"""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Literal

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from src.agent.tools._report_screenshot import convert_to_pdf, render_pdf_page_with_footer

COLUMNS = [
    "序号", "数据类别", "明细分类", "金额区间",
    "下单笔数", "成功笔数", "下单金额", "成功金额", "成功率", "备注",
]
N_COLS = len(COLUMNS)

TITLE_ROW = 1
HEADER_ROW = 2
DATA_FIRST_ROW = 3

SUCCESS_STATUS = "成功"

# (标签, 上界含/None=无上界)，区间连续——每一档下界就是上一档的上界（严格大于），不
# 单独存下界，避免像 (0,500)/(501,2000) 那样在 500~501 之间留出空隙导致 500.5 这类
# 小数金额无法命中任何一档。顺序固定，不受当次数据影响。
AMOUNT_BANDS: list[tuple[str, float | None]] = [
    ("0-500", 500),
    ("501-2000", 2000),
    ("2001-10000", 10000),
    ("10001-30000", 30000),
    ("30001以上", None),
]

_TITLE_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
_TITLE_FONT = Font(color="FFFFFF", bold=True, size=14)
_HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_HEADER_FONT = Font(bold=True)
_SUBTOTAL_FILL = PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid")
_TOTAL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

_AMOUNT_FORMAT = "#,##0.00"
_PERCENT_FORMAT = "0.0%"

# 固定手工列宽——不用 excel_tools._autosize：第 1 行是跨 10 列的合并长标题，喂给
# 按列独立测宽的 autosize 会把 A 列撑得离谱，跟 A 列实际内容（1~2 位数的序号）无关。
_COLUMN_WIDTHS = {1: 6, 2: 18, 3: 18, 4: 14, 5: 10, 6: 10, 7: 16, 8: 16, 9: 10, 10: 22}


@dataclass
class Agg:
    count: int
    success_count: int
    amount: float
    success_amount: float

    @property
    def rate(self) -> float | None:
        return self.success_count / self.count if self.count else None


@dataclass
class DataRow:
    row_kind: Literal["band", "subtotal", "total"]
    data_category: str
    # band/subtotal 行：该行所属明细分类的原文本；total 行：f"{data_category}合计"
    detail_category_label: str
    band_label: str  # "0-500" 等区间标签 / "小计" / "合计"
    count: int
    success_count: int
    amount: float
    success_amount: float
    rate: float | None
    remark: str | None = None


@dataclass
class MergeSpan:
    start: int  # plan.rows 下标（含）
    end: int  # plan.rows 下标（含）
    value: str


@dataclass
class SheetPlan:
    rows: list[DataRow] = field(default_factory=list)
    data_category_spans: list[MergeSpan] = field(default_factory=list)
    detail_category_spans: list[MergeSpan] = field(default_factory=list)


def band_label(order_amount: float) -> str:
    for label, high in AMOUNT_BANDS:
        if high is None or order_amount <= high:
            return label
    return AMOUNT_BANDS[-1][0]


def _aggregate(records: list[dict]) -> Agg:
    success = [r for r in records if r["status"] == SUCCESS_STATUS]
    return Agg(
        count=len(records),
        success_count=len(success),
        amount=sum(r["order_amount"] for r in records),
        success_amount=sum(r["order_amount"] for r in success),
    )


def group_records(records: list[dict]) -> dict[str, dict[str, list[dict]]]:
    """按 data_category -> detail_category 两级分组。

    普通 dict 的插入顺序即首次出现顺序，天然实现"类别顺序完全由数据决定"，不需要
    任何排序或预先配置的分类清单。
    """
    grouped: dict[str, dict[str, list[dict]]] = {}
    for rec in records:
        grouped.setdefault(rec["data_category"], {}).setdefault(rec["detail_category"], []).append(rec)
    return grouped


def build_sheet_plan(records: list[dict]) -> SheetPlan:
    grouped = group_records(records)
    plan = SheetPlan()

    for data_category, detail_map in grouped.items():
        dc_start = len(plan.rows)
        dc_all_records: list[dict] = []

        for detail_category, detail_records in detail_map.items():
            dc_all_records.extend(detail_records)
            detail_start = len(plan.rows)

            buckets: dict[str, list[dict]] = {label: [] for label, _ in AMOUNT_BANDS}
            for r in detail_records:
                buckets[band_label(r["order_amount"])].append(r)

            for label, _ in AMOUNT_BANDS:
                bucket = buckets[label]
                if not bucket:
                    continue  # 该区间当前没有记录：跳过，不出零值行
                agg = _aggregate(bucket)
                plan.rows.append(
                    DataRow(
                        row_kind="band",
                        data_category=data_category,
                        detail_category_label=detail_category,
                        band_label=label,
                        count=agg.count,
                        success_count=agg.success_count,
                        amount=agg.amount,
                        success_amount=agg.success_amount,
                        rate=agg.rate,
                    )
                )

            sub_agg = _aggregate(detail_records)  # 全量汇总，数值上等于可见区间行之和
            plan.rows.append(
                DataRow(
                    row_kind="subtotal",
                    data_category=data_category,
                    detail_category_label=detail_category,
                    band_label="小计",
                    count=sub_agg.count,
                    success_count=sub_agg.success_count,
                    amount=sub_agg.amount,
                    success_amount=sub_agg.success_amount,
                    rate=sub_agg.rate,
                )
            )

            detail_end = len(plan.rows) - 1
            plan.detail_category_spans.append(MergeSpan(detail_start, detail_end, detail_category))

        total_agg = _aggregate(dc_all_records)  # 该数据类别下所有明细分类简单相加
        plan.rows.append(
            DataRow(
                row_kind="total",
                data_category=data_category,
                detail_category_label=f"{data_category}合计",
                band_label="合计",
                count=total_agg.count,
                success_count=total_agg.success_count,
                amount=total_agg.amount,
                success_amount=total_agg.success_amount,
                rate=total_agg.rate,
            )
        )

        dc_end = len(plan.rows) - 1  # 含合计行本身，合计行视觉上仍属于该数据类别
        plan.data_category_spans.append(MergeSpan(dc_start, dc_end, data_category))

    return plan


def derive_title_text(report_date: str) -> str:
    """标题里的日期直接取接口返回的 date 字段（ISO 格式，代表这批数据是哪一天的），
    不再从每条记录的 occurred_at 反推——接口现在已经明确给出这批数据对应的自然日。"""
    d = date.fromisoformat(report_date)
    return f"支付宝匹配数据表（{d.month}月{d.day}日）整日数据"


def render_workbook(plan: SheetPlan, title_text: str, out_path: Path) -> None:
    wb = Workbook()
    ws = wb.worksheets[0]  # 用 worksheets[0] 而不是 wb.active——后者的类型标注是
    # Worksheet | None，会让下面每一处 ws.xxx 调用都被类型检查器误报；新建的
    # Workbook() 必定有且只有这一张默认 sheet，worksheets[0] 的类型标注是非 None 的。
    ws.title = "支付宝匹配数据表"

    ws.cell(row=TITLE_ROW, column=1, value=title_text)
    ws.merge_cells(start_row=TITLE_ROW, start_column=1, end_row=TITLE_ROW, end_column=N_COLS)
    title_cell = ws.cell(row=TITLE_ROW, column=1)
    title_cell.fill = _TITLE_FILL
    title_cell.font = _TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[TITLE_ROW].height = 26

    for col_idx, name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx, value=name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, row in enumerate(plan.rows):
        r = DATA_FIRST_ROW + i
        ws.cell(row=r, column=1, value=i + 1)  # 序号：全表连续计数，不合并
        ws.cell(row=r, column=4, value=row.band_label)
        ws.cell(row=r, column=5, value=row.count)
        ws.cell(row=r, column=6, value=row.success_count)
        c7 = ws.cell(row=r, column=7, value=row.amount)
        c7.number_format = _AMOUNT_FORMAT
        c8 = ws.cell(row=r, column=8, value=row.success_amount)
        c8.number_format = _AMOUNT_FORMAT
        c9 = ws.cell(row=r, column=9, value=row.rate)
        c9.number_format = _PERCENT_FORMAT
        ws.cell(row=r, column=10, value=row.remark)

        if row.row_kind == "total":
            ws.cell(row=r, column=3, value=row.detail_category_label)  # 合计行独立写，不参与合并

        fill = _SUBTOTAL_FILL if row.row_kind == "subtotal" else _TOTAL_FILL if row.row_kind == "total" else None
        if fill:
            for c in range(1, N_COLS + 1):
                ws.cell(row=r, column=c).fill = fill

    for span in plan.data_category_spans:
        top, bottom = DATA_FIRST_ROW + span.start, DATA_FIRST_ROW + span.end
        ws.cell(row=top, column=2, value=span.value)
        if bottom > top:
            ws.merge_cells(start_row=top, start_column=2, end_row=bottom, end_column=2)
        ws.cell(row=top, column=2).alignment = Alignment(horizontal="center", vertical="center")

    for span in plan.detail_category_spans:
        top, bottom = DATA_FIRST_ROW + span.start, DATA_FIRST_ROW + span.end
        ws.cell(row=top, column=3, value=span.value)
        if bottom > top:
            ws.merge_cells(start_row=top, start_column=3, end_row=bottom, end_column=3)
        ws.cell(row=top, column=3).alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, width in _COLUMN_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = f"A{DATA_FIRST_ROW}"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def render_screenshot(xlsx_path: Path, out_path: Path, footer_text: str, tmp_dir: Path) -> None:
    """把 render_workbook 生成的单一 sheet 原样截图成一张 PNG，贴上 footer_text 脚注。

    只有一张 sheet、没有公式、没有跨 sheet 引用——不需要像成本报表那样先拆 sheet/摊平
    公式，直接在整张已用区域上设 print_area + 强制单页缩放即可转 PDF 截图。
    """
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.worksheets[0]
    ws.print_area = f"A1:{get_column_letter(N_COLS)}{ws.max_row}"
    ws.print_options.gridLines = False
    ws.print_options.headings = False
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    ws.page_margins = PageMargins(left=0.2, right=0.2, top=0.2, bottom=0.2, header=0, footer=0)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    src_path = tmp_dir / "alipay_screenshot_src.xlsx"
    wb.save(src_path)

    pdf_outdir = tmp_dir / "pdf_out"
    pdf_profile = tmp_dir / "pdf_profile"
    pdf_outdir.mkdir(exist_ok=True)
    pdf_profile.mkdir(exist_ok=True)
    pdf_path = convert_to_pdf(src_path, pdf_outdir, pdf_profile)

    render_pdf_page_with_footer(pdf_path, 0, out_path, footer_text)
