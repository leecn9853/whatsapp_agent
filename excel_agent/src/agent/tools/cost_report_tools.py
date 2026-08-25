"""成本报表 skill 的工具：调用模拟第三方数据服务（thrid_app）拉取最新采购/成本数据，填进固定
模板 Excel（templates/cost_report.xlsx），用 LibreOffice headless 重算公式后，把「成本分析看板」
里的一个区块渲染成表格图或图表图片发给用户。

模板结构（sheet 名、表头列顺序、单元格坐标、看板区块行列范围）是写死的，不是通用 Excel 处理
场景，所以这里只暴露一个确定性的原子工具 generate_cost_report_image，内部一次性完成"拉两个
接口 -> 填模板 -> LibreOffice 重算 -> 读看板区块 -> 画图 -> 返回 PNG"。

同一个任务（同一个 report_id，即 agent_server runs_store 的 run_id）如果先后调用了 table 和
chart 两种 render_type，第二次调用会复用第一次已经拉好、重算好的数据（见快照目录里的
recalculated.xlsx），不会重复拉第三方接口、不会重复跑一遍 LibreOffice——保证同一个任务里两张
图看到的是同一份数据。不同任务（不同 report_id）之间不共享，下一个任务会重新拉取最新数据。

不再生成/发送 Excel 文件给用户——Excel 只是内部中间产物，最终产物是 PNG 图片。
"""
from __future__ import annotations

import os
import shutil
import tempfile
from datetime import date
from pathlib import Path
from typing import Literal

import httpx
import openpyxl
from langchain.tools import ToolRuntime
from langchain_core.tools import tool

from src.context import ContextSchema
from src.agent.tools._cost_report_render import (
    render_chart_screenshot,
    render_dashboard_screenshot,
)
from src.agent.tools._naming import build_stem
from src.agent.tools._report_screenshot import ReportRenderError, recalculate_with_libreoffice
from src.agent.tools._report_snapshot import resolve_snapshot_dir
from src.agent.tools.excel_tools import OUTPUT_DIR, _user_dir

THIRD_APP_BASE_URL = os.getenv("THIRD_APP_BASE_URL", "http://127.0.0.1:8800")

_PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent.parent
TEMPLATE_PATH = _PROJECT_ROOT / "templates" / "cost_report.xlsx"

# 快照目录（按用户再按任务分两级文件夹）的解析逻辑见 _report_snapshot.py（被 alipay 报表
# 共用）。同一个任务里 table/chart 两次调用共享同一份 recalculated.xlsx（见
# generate_cost_report_image 里的复用逻辑），不会各自留一份。

# files.py 用这个集合判断哪些工具调用产生的文件需要自动发给用户（返回值是文件路径）。
COST_REPORT_OUTPUT_TOOL_NAMES = {"generate_cost_report_image"}

_MONTHLY_COST_SHEET = "月度成本明细"
_MONTHLY_COST_FIELDS = ["month", "department", "cost_category", "amount", "budget", "over_budget"]

_SUPPLIER_SHEET = "供应商采购成本"
_SUPPLIER_FIELDS = [
    "supplier_name",
    "purchase_category",
    "purchase_date",
    "purchase_amount",
    "payment_status",
]

_MAX_PAGES = 20  # 防止 total 异常返回导致死循环的保护上限

SECTIONS: dict[str, dict] = {
    "monthly_trend": {
        "title": "月度成本趋势",
        "data_first_row": 5,
        "total_row": 17,
        "columns": ["月份", "总支出(万元)", "总预算(万元)", "差异(万元)", "预算执行率", "超支占比", "超支笔数", "总笔数"],
        "percent_cols": ["预算执行率", "超支占比"],
        "sort_by": "月份",
        "chart": {"kind": "line", "x": "月份", "series": ["总支出(万元)", "总预算(万元)"]},
    },
    "department_budget": {
        "title": "部门预算执行对比",
        "data_first_row": 22,
        "total_row": 28,
        "columns": ["部门", "年度总成本(万元)", "年度预算(万元)", "差异(万元)"],
        "percent_cols": [],
        "chart": {"kind": "bar", "x": "部门", "series": ["年度总成本(万元)", "年度预算(万元)"]},
    },
    "purchase_category": {
        "title": "采购类别成本对比",
        "data_first_row": 33,
        "total_row": 39,
        "columns": [
            "成本类别",
            "供应商采购总额(万元)",
            "明细记录总额(万元)",
            "供应商采购占比",
            "待付款笔数占比",
            "待付款笔数",
            "总笔数",
        ],
        "percent_cols": ["供应商采购占比", "待付款笔数占比"],
        "chart": {"kind": "bar", "x": "成本类别", "series": ["供应商采购总额(万元)", "明细记录总额(万元)"]},
    },
}


def _fetch_supplier_purchases(client: httpx.Client) -> list[dict]:
    resp = client.get(f"{THIRD_APP_BASE_URL}/api/electronics/orders")
    resp.raise_for_status()
    return resp.json()


def _fetch_monthly_costs(client: httpx.Client) -> list[dict]:
    page_size = 99
    items: list[dict] = []
    page = 1
    total = None
    while (total is None or len(items) < total) and page <= _MAX_PAGES:
        resp = client.get(
            f"{THIRD_APP_BASE_URL}/api/food-agri/orders",
            params={"page": page, "page_size": page_size},
        )
        resp.raise_for_status()
        data = resp.json()
        total = data["total"]
        items.extend(data["items"])
        if not data["items"]:
            break
        page += 1
    return items


def _clear_data_rows(ws, n_cols: int) -> None:
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=n_cols):
        for cell in row:
            cell.value = None


def _write_rows(ws, records: list[dict], fields: list[str]) -> None:
    for i, record in enumerate(records):
        row = i + 2
        for j, field in enumerate(fields):
            col = j + 1
            value = record[field]
            if field == "purchase_date":
                value = date.fromisoformat(value)
            ws.cell(row=row, column=col, value=value)


def _build_filled_workbook(client: httpx.Client, dest_path: Path) -> tuple[list[dict], list[dict]]:
    """拉最新数据，写进 dest_path 处的模板拷贝里的两张原始明细 sheet，返回拉到的数据。"""
    supplier_rows = _fetch_supplier_purchases(client)
    monthly_rows = _fetch_monthly_costs(client)

    dest_path.write_bytes(TEMPLATE_PATH.read_bytes())
    wb = openpyxl.load_workbook(dest_path)

    monthly_ws = wb[_MONTHLY_COST_SHEET]
    _clear_data_rows(monthly_ws, len(_MONTHLY_COST_FIELDS))
    _write_rows(monthly_ws, monthly_rows, _MONTHLY_COST_FIELDS)

    supplier_ws = wb[_SUPPLIER_SHEET]
    _clear_data_rows(supplier_ws, len(_SUPPLIER_FIELDS))
    _write_rows(supplier_ws, supplier_rows, _SUPPLIER_FIELDS)

    wb.save(dest_path)
    return monthly_rows, supplier_rows


def _count_detail_rows(recalculated_path: Path) -> tuple[int, int]:
    """同一个任务里复用已有 recalculated.xlsx 时，跳过了重新拉接口，改从明细 sheet 数据行数
    还原「拉到 N 条」的条数，给用户看的提示文案用（跟当次真的拉接口时口径一致）。
    """
    wb = openpyxl.load_workbook(recalculated_path, data_only=True)
    monthly_count = sum(1 for row in wb[_MONTHLY_COST_SHEET].iter_rows(min_row=2, max_col=1) if row[0].value is not None)
    supplier_count = sum(1 for row in wb[_SUPPLIER_SHEET].iter_rows(min_row=2, max_col=1) if row[0].value is not None)
    return monthly_count, supplier_count


def _resolve_image_save_path(render_type: str, ctx) -> Path:
    user_output_dir = _user_dir(OUTPUT_DIR, ctx)
    stem = build_stem(f"成本报表_{render_type}", ctx)
    candidate = user_output_dir / f"{stem}.png"
    n = 1
    while candidate.exists():
        candidate = user_output_dir / f"{stem}({n}).png"
        n += 1
    return candidate


@tool(response_format="content_and_artifact")
def generate_cost_report_image(
    runtime: ToolRuntime[ContextSchema],
    render_type: Literal["table", "chart"] = "chart",
) -> tuple[str, str | None]:
    """生成成本报表图片：拉取最新数据、套用模板、渲染成一张 PNG 截图，一次调用完成整个流程。

    render_type: "chart"（默认，图表形式）｜"table"（表格明细/合计形式，用户要具体数字时用）。
    参数选择依据和报表内容说明见 cost-report 技能的 SKILL.md。

    仅当用户明确要生成/查看成本报表（表格或图表）时才调用。
    """
    try:
        report_id, snapshot_dir = resolve_snapshot_dir(runtime.context)
        shared_recalculated_path = snapshot_dir / "recalculated.xlsx"

        with tempfile.TemporaryDirectory(prefix="cost_report_") as tmp_root:
            tmp_root_path = Path(tmp_root)

            if shared_recalculated_path.exists():
                # 同一个任务（同一个 report_id）已经有另一次 render_type 调用拉过数据、重算
                # 过了——直接复用那份结果，不重复拉第三方接口、不重复跑 LibreOffice，保证
                # 同一个任务里 table/chart 两张图看到的是同一份数据。
                recalculated_path = shared_recalculated_path
                monthly_count, supplier_count = _count_detail_rows(recalculated_path)
            else:
                with httpx.Client(timeout=30.0) as client:
                    filled_path = tmp_root_path / "filled.xlsx"
                    monthly_rows, supplier_rows = _build_filled_workbook(client, filled_path)
                monthly_count, supplier_count = len(monthly_rows), len(supplier_rows)

                profile_dir = tmp_root_path / "lo_profile"
                outdir = tmp_root_path / "lo_out"
                profile_dir.mkdir()
                outdir.mkdir()
                recalculated_path = recalculate_with_libreoffice(filled_path, outdir, profile_dir)
                shutil.copy2(recalculated_path, shared_recalculated_path)

            footer_text = f"报表ID: {report_id}"
            save_path = _resolve_image_save_path(render_type, runtime.context)

            if render_type == "table":
                render_dashboard_screenshot(
                    recalculated_path, list(SECTIONS.values()), save_path, footer_text, tmp_root_path
                )
            else:
                render_chart_screenshot(recalculated_path, save_path, footer_text, tmp_root_path)

            # 最终发给用户的图片也备份一份进快照文件夹，方便跟共享的 recalculated.xlsx 对照；
            # 两种 render_type 各自的图片都保留，只有底层数据快照是共享的那一份。
            shutil.copy2(save_path, snapshot_dir / save_path.name)
    except httpx.HTTPError as e:
        return (
            f"拉取成本数据失败：{e}\n"
            f"请确认模拟数据服务已启动且 THIRD_APP_BASE_URL（当前为 {THIRD_APP_BASE_URL!r}）配置正确。",
            None,
        )
    except ReportRenderError as e:
        return (str(e), None)

    sheet_name = "成本分析看板" if render_type == "table" else "成本图表"
    kind = "表格截图" if render_type == "table" else "图表截图"
    content = (
        f"已拉取最新的月度成本明细（{monthly_count} 条）和供应商采购数据（{supplier_count} 条），"
        f"生成「{sheet_name}」整体{kind}：{save_path.name}"
    )
    return content, str(save_path)
