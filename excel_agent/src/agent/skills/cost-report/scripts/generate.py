#!/usr/bin/env python3
"""成本报表图片生成 —— 沙箱容器内跑的 CLI 版本，供 deepagents 内置的 execute 工具调用。

跟原来 src/agent/tools/cost_report_tools.py（@tool 版本，已废弃/待删）流程完全一致：拉最新
数据 -> 填模板 -> LibreOffice 重算公式 -> 把成本分析看板/成本图表整体截图成 PNG。区别只是
参数从 Python 对象（ContextSchema）换成命令行字符串参数，且本文件是自包含的——不 import
项目源码（容器里也没有那份代码，只有 skills/ 这一份 bind mount），LibreOffice 截图原语
（ReportRenderError/_soffice_convert 等）、命名/快照目录约定都在本文件内重新实现一份。

输出契约（调用方靠这个判断成功/失败，不看 exit_code 之外的东西）：
- 成功：最后一行 stdout 打印 `RESULT_PATH:<容器内绝对路径>`，exit 0。
- 已知失败（拉数据失败、LibreOffice 不可用/转换失败）：最后一行 stdout 打印
  `ERROR:<中文说明>`，exit 1。LibreOffice 自己往 stderr 打的噪声不会污染这行。
- 未预期的异常：完整 traceback 打到 stderr（供开发排查），stdout 最后一行仍是简化的
  `ERROR:脚本内部错误：<e>`，exit 1。
"""
from __future__ import annotations

import argparse
import io
import os
import random
import re
import shutil
import subprocess
import sys
import tempfile
import traceback
from datetime import date, datetime
from pathlib import Path
from typing import Literal

import httpx
import openpyxl
import pymupdf  # 把 LibreOffice 转出来的 PDF 页面渲成图片
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from PIL import Image, ImageDraw, ImageFont

SCRIPT_DIR = Path(__file__).resolve().parent
SKILL_DIR = SCRIPT_DIR.parent
# 容器里固定是 /workspace/output|snapshots；env 覆盖只用于宿主机本地验证（见
# docs/skills-tools-refactor-plan.md 阶段二 Stage 3 验证步骤 1），不作为对外契约。
OUTPUT_DIR = Path(os.getenv("COST_REPORT_OUTPUT_DIR", "/workspace/output"))
SNAPSHOT_ROOT = Path(os.getenv("COST_REPORT_SNAPSHOT_ROOT", "/workspace/snapshots"))
TEMPLATE_PATH = SKILL_DIR / "template" / "cost_report.xlsx"
FONT_PATH = SKILL_DIR / "assets" / "NotoSansSC-Regular.ttf"

SOFFICE_BIN = os.getenv("SOFFICE_BIN", "soffice")
THIRD_APP_BASE_URL = os.getenv("THIRD_APP_BASE_URL", "http://host.docker.internal:8800")

_MONTHLY_COST_SHEET = "月度成本明细"
_MONTHLY_COST_FIELDS = ["month", "department", "cost_category", "amount", "budget", "over_budget"]

_SUPPLIER_SHEET = "供应商采购成本"
_SUPPLIER_FIELDS = [
    "supplier_name",
    "purchase_category",
    "purchase_date",
    "purchase_amount",
    "payment_status",
]

_DASHBOARD_SHEET = "成本分析看板"
_CHART_SHEET = "成本图表"

_MAX_PAGES = 20  # 防止 total 异常返回导致死循环的保护上限

SECTIONS: dict[str, dict] = {
    "monthly_trend": {
        "title": "月度成本趋势",
        "data_first_row": 5,
        "total_row": 17,
        "columns": ["月份", "总支出(万元)", "总预算(万元)", "差异(万元)", "预算执行率", "超支占比", "超支笔数", "总笔数"],
    },
    "department_budget": {
        "title": "部门预算执行对比",
        "data_first_row": 22,
        "total_row": 28,
        "columns": ["部门", "年度总成本(万元)", "年度预算(万元)", "差异(万元)"],
    },
    "purchase_category": {
        "title": "采购类别成本对比",
        "data_first_row": 33,
        "total_row": 39,
        "columns": [
            "成本类别",
            "供应商采购总额(万元)",
            "明细记录总额(万元)",
            "供应商采购占比",
            "待付款笔数占比",
            "待付款笔数",
            "总笔数",
        ],
    },
}


class ReportRenderError(Exception):
    """渲染/重算阶段的内部错误，转成 ERROR: 输出，不让异常裸露到顶层。"""


# ---------------------------------------------------------------------------
# 命名/快照目录约定（_naming.py + _report_snapshot.py 的字符串参数版本）
# ---------------------------------------------------------------------------

_UNSAFE_CHARS = re.compile(r'[\\/:*?"<>|\s]+')


def sanitize(text: str, max_len: int = 50) -> str:
    safe = _UNSAFE_CHARS.sub("_", text).strip("_")
    return safe[:max_len] or "untitled"


def sanitize_user_id(user_id: str) -> str:
    return sanitize(str(user_id).split("@")[0])


def build_stem(title_stem: str, caller: str, user_id: str | None) -> str:
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    random_suffix = f"{random.randint(0, 999999):06d}"
    safe_title = sanitize(title_stem)
    if caller == "whatsapp" and user_id:
        return f"{safe_title}_{timestamp}_{sanitize_user_id(user_id)}_{random_suffix}"
    return f"DEBUG_{safe_title}_{timestamp}_{random_suffix}"


def _snapshot_user_folder(caller: str, user_id: str | None) -> str:
    if user_id:
        return sanitize_user_id(user_id)
    return caller


def resolve_snapshot_dir(report_id: str, caller: str, user_id: str | None) -> Path:
    snapshot_dir = SNAPSHOT_ROOT / _snapshot_user_folder(caller, user_id) / report_id
    snapshot_dir.mkdir(parents=True, exist_ok=True)
    return snapshot_dir


def _resolve_output_dir(caller: str, user_id: str | None) -> Path:
    if caller == "whatsapp" and user_id:
        d = OUTPUT_DIR / sanitize_user_id(user_id)
        d.mkdir(parents=True, exist_ok=True)
        return d
    return OUTPUT_DIR


def _resolve_image_save_path(render_type: str, caller: str, user_id: str | None) -> Path:
    user_output_dir = _resolve_output_dir(caller, user_id)
    stem = build_stem(f"成本报表_{render_type}", caller, user_id)
    candidate = user_output_dir / f"{stem}.png"
    n = 1
    while candidate.exists():
        candidate = user_output_dir / f"{stem}({n}).png"
        n += 1
    return candidate


# ---------------------------------------------------------------------------
# LibreOffice/pymupdf 截图原语（_report_screenshot.py 原样搬）
# ---------------------------------------------------------------------------


def _soffice_convert(input_path: Path, outdir: Path, profile_dir: Path, fmt: str) -> Path:
    """outdir/profile_dir 必须每次调用用全新目录（调用方靠 tempfile.TemporaryDirectory 保证），
    避免并发调用抢同一份 LibreOffice 用户配置锁——这条不变量不能在迁移时弄丢。
    """
    try:
        result = subprocess.run(
            [
                SOFFICE_BIN,
                "--headless",
                "--calc",
                "--convert-to",
                fmt,
                "--outdir",
                str(outdir),
                str(input_path),
                f"-env:UserInstallation=file://{profile_dir}",
            ],
            capture_output=True,
            text=True,
            timeout=60,
        )
    except FileNotFoundError as e:
        raise ReportRenderError(
            "未找到 soffice 命令，请确认容器已安装 LibreOffice。"
        ) from e
    except subprocess.TimeoutExpired as e:
        raise ReportRenderError("LibreOffice 转换超时。") from e

    if result.returncode != 0:
        raise ReportRenderError(f"LibreOffice 转换失败：{result.stderr or result.stdout}")

    converted = outdir / f"{input_path.stem}.{fmt}"
    if not converted.exists():
        raise ReportRenderError("LibreOffice 没有生成转换后的文件。")
    return converted


def recalculate_with_libreoffice(input_path: Path, outdir: Path, profile_dir: Path) -> Path:
    return _soffice_convert(input_path, outdir, profile_dir, "xlsx")


def convert_to_pdf(input_path: Path, outdir: Path, profile_dir: Path) -> Path:
    return _soffice_convert(input_path, outdir, profile_dir, "pdf")


def render_pdf_page_with_footer(pdf_path: Path, page_index: int, out_path: Path, footer_text: str) -> None:
    doc = pymupdf.open(pdf_path)
    pix = doc[page_index].get_pixmap(dpi=200)
    doc.close()
    section_img = Image.open(io.BytesIO(pix.tobytes("png"))).convert("RGB")

    footer_height = 28
    canvas = Image.new("RGB", (section_img.width, section_img.height + footer_height), "white")
    canvas.paste(section_img, (0, 0))
    draw = ImageDraw.Draw(canvas)
    font = ImageFont.truetype(str(FONT_PATH), 16)
    text_width = draw.textlength(footer_text, font=font)
    draw.text(
        ((section_img.width - text_width) / 2, section_img.height + 4),
        footer_text,
        font=font,
        fill="#666666",
    )
    canvas.save(out_path)


# ---------------------------------------------------------------------------
# 数据拉取 + 填模板（cost_report_tools.py 原样搬）
# ---------------------------------------------------------------------------


def _fetch_supplier_purchases(client: httpx.Client) -> list[dict]:
    resp = client.get(f"{THIRD_APP_BASE_URL}/api/electronics/orders")
    resp.raise_for_status()
    return resp.json()


def _fetch_monthly_costs(client: httpx.Client) -> list[dict]:
    page_size = 99
    items: list[dict] = []
    page = 1
    total = None
    while (total is None or len(items) < total) and page <= _MAX_PAGES:
        resp = client.get(
            f"{THIRD_APP_BASE_URL}/api/food-agri/orders",
            params={"page": page, "page_size": page_size},
        )
        resp.raise_for_status()
        data = resp.json()
        total = data["total"]
        items.extend(data["items"])
        if not data["items"]:
            break
        page += 1
    return items


def _clear_data_rows(ws, n_cols: int) -> None:
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=n_cols):
        for cell in row:
            cell.value = None


def _write_rows(ws, records: list[dict], fields: list[str]) -> None:
    for i, record in enumerate(records):
        row = i + 2
        for j, field in enumerate(fields):
            col = j + 1
            value = record[field]
            if field == "purchase_date":
                value = date.fromisoformat(value)
            ws.cell(row=row, column=col, value=value)


def _build_filled_workbook(client: httpx.Client, dest_path: Path) -> tuple[list[dict], list[dict]]:
    supplier_rows = _fetch_supplier_purchases(client)
    monthly_rows = _fetch_monthly_costs(client)

    dest_path.write_bytes(TEMPLATE_PATH.read_bytes())
    wb = openpyxl.load_workbook(dest_path)

    monthly_ws = wb[_MONTHLY_COST_SHEET]
    _clear_data_rows(monthly_ws, len(_MONTHLY_COST_FIELDS))
    _write_rows(monthly_ws, monthly_rows, _MONTHLY_COST_FIELDS)

    supplier_ws = wb[_SUPPLIER_SHEET]
    _clear_data_rows(supplier_ws, len(_SUPPLIER_FIELDS))
    _write_rows(supplier_ws, supplier_rows, _SUPPLIER_FIELDS)

    wb.save(dest_path)
    return monthly_rows, supplier_rows


def _count_detail_rows(recalculated_path: Path) -> tuple[int, int]:
    wb = openpyxl.load_workbook(recalculated_path, data_only=True)
    monthly_count = sum(1 for row in wb[_MONTHLY_COST_SHEET].iter_rows(min_row=2, max_col=1) if row[0].value is not None)
    supplier_count = sum(1 for row in wb[_SUPPLIER_SHEET].iter_rows(min_row=2, max_col=1) if row[0].value is not None)
    return monthly_count, supplier_count


# ---------------------------------------------------------------------------
# 看板/图表整体截图（_cost_report_render.py 里截图路径原样搬，matplotlib 重画路径
# 原工具已经不用，这里不再迁移）
# ---------------------------------------------------------------------------


def _content_width_units(value) -> float:
    text = "" if value is None else str(value)
    return sum(1.8 if ord(ch) > 0x2E80 else 1.0 for ch in text)


def _flatten_dashboard_formulas(wb, wb_values) -> None:
    ws_values = wb_values[_DASHBOARD_SHEET]
    ws = wb[_DASHBOARD_SHEET]
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.value = ws_values.cell(row=cell.row, column=cell.column).value


def render_dashboard_screenshot(
    recalculated_path: Path, sections_cfg: list[dict], out_path: Path, footer_text: str, tmp_dir: Path
) -> None:
    wb_values = openpyxl.load_workbook(recalculated_path, data_only=True)
    wb = openpyxl.load_workbook(recalculated_path)

    ws = wb[_DASHBOARD_SHEET]
    _flatten_dashboard_formulas(wb, wb_values)

    for name in list(wb.sheetnames):
        if name != _DASHBOARD_SHEET:
            del wb[name]

    max_n_cols = max(len(cfg["columns"]) for cfg in sections_cfg)

    blocks = []
    for cfg in sections_cfg:
        title_row = cfg["data_first_row"] - 2
        n_cols = len(cfg["columns"])
        blocks.append(
            {
                "title_row": title_row,
                "header_row": title_row + 1,
                "total_row": cfg["total_row"],
                "n_cols": n_cols,
            }
        )

    for col_idx in range(1, max_n_cols + 1):
        widths = [
            _content_width_units(ws.cell(row=r, column=col_idx).value)
            for block in blocks
            for r in range(block["header_row"], block["total_row"] + 1)
        ]
        ws.column_dimensions[get_column_letter(col_idx)].width = max(widths, default=8.0) + 3

    for block in blocks:
        n_cols = block["n_cols"]
        first_col, last_col = 1, n_cols
        actual_title = ws.cell(row=block["title_row"], column=first_col).value
        title_units = _content_width_units(actual_title) * 1.2
        total_width = sum(ws.column_dimensions[get_column_letter(c)].width for c in range(first_col, last_col + 1))
        if title_units > total_width:
            extra_per_col = (title_units - total_width) / n_cols
            for c in range(first_col, last_col + 1):
                ws.column_dimensions[get_column_letter(c)].width += extra_per_col
        ws.merge_cells(f"{get_column_letter(first_col)}{block['title_row']}:{get_column_letter(last_col)}{block['title_row']}")

    top_row = blocks[0]["title_row"]
    bottom_row = blocks[-1]["total_row"]
    ws.print_area = f"A{top_row}:{get_column_letter(max_n_cols)}{bottom_row}"
    ws.print_options.gridLines = False
    ws.print_options.headings = False
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    ws.page_margins = PageMargins(left=0.2, right=0.2, top=0.2, bottom=0.2, header=0, footer=0)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    src_path = tmp_dir / "dashboard_screenshot_src.xlsx"
    wb.save(src_path)

    pdf_outdir = tmp_dir / "pdf_out"
    pdf_profile = tmp_dir / "pdf_profile"
    pdf_outdir.mkdir(exist_ok=True)
    pdf_profile.mkdir(exist_ok=True)
    pdf_path = convert_to_pdf(src_path, pdf_outdir, pdf_profile)

    render_pdf_page_with_footer(pdf_path, 0, out_path, footer_text)


def render_chart_screenshot(recalculated_path: Path, out_path: Path, footer_text: str, tmp_dir: Path) -> None:
    wb_values = openpyxl.load_workbook(recalculated_path, data_only=True)
    wb = openpyxl.load_workbook(recalculated_path)

    _flatten_dashboard_formulas(wb, wb_values)

    for name in list(wb.sheetnames):
        if name not in (_DASHBOARD_SHEET, _CHART_SHEET):
            del wb[name]

    wb[_DASHBOARD_SHEET].sheet_state = "hidden"

    ws = wb[_CHART_SHEET]
    ws.print_area = "A1:M56"
    ws.print_options.gridLines = False
    ws.print_options.headings = False
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    ws.page_margins = PageMargins(left=0.2, right=0.2, top=0.2, bottom=0.2, header=0, footer=0)
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    src_path = tmp_dir / "chart_screenshot_src.xlsx"
    wb.save(src_path)

    pdf_outdir = tmp_dir / "pdf_out"
    pdf_profile = tmp_dir / "pdf_profile"
    pdf_outdir.mkdir(exist_ok=True)
    pdf_profile.mkdir(exist_ok=True)
    pdf_path = convert_to_pdf(src_path, pdf_outdir, pdf_profile)

    doc = pymupdf.open(pdf_path)
    last_page = len(doc) - 1
    doc.close()
    render_pdf_page_with_footer(pdf_path, last_page, out_path, footer_text)


# ---------------------------------------------------------------------------
# 主流程
# ---------------------------------------------------------------------------


def generate(render_type: Literal["table", "chart"], report_id: str, caller: str, user_id: str | None) -> Path:
    snapshot_dir = resolve_snapshot_dir(report_id, caller, user_id)
    shared_recalculated_path = snapshot_dir / "recalculated.xlsx"

    with tempfile.TemporaryDirectory(prefix="cost_report_") as tmp_root:
        tmp_root_path = Path(tmp_root)

        if shared_recalculated_path.exists():
            # 同一个任务（同一个 report_id）已经有另一次 render_type 调用拉过数据、重算过
            # 了——直接复用，不重复拉第三方接口、不重复跑 LibreOffice。
            recalculated_path = shared_recalculated_path
        else:
            with httpx.Client(timeout=30.0) as client:
                filled_path = tmp_root_path / "filled.xlsx"
                _build_filled_workbook(client, filled_path)

            profile_dir = tmp_root_path / "lo_profile"
            outdir = tmp_root_path / "lo_out"
            profile_dir.mkdir()
            outdir.mkdir()
            recalculated_path = recalculate_with_libreoffice(filled_path, outdir, profile_dir)
            shutil.copy2(recalculated_path, shared_recalculated_path)

        footer_text = f"报表ID: {report_id}"
        save_path = _resolve_image_save_path(render_type, caller, user_id)

        if render_type == "table":
            render_dashboard_screenshot(
                recalculated_path, list(SECTIONS.values()), save_path, footer_text, tmp_root_path
            )
        else:
            render_chart_screenshot(recalculated_path, save_path, footer_text, tmp_root_path)

        shutil.copy2(save_path, snapshot_dir / save_path.name)

    return save_path


def main() -> int:
    parser = argparse.ArgumentParser(description="生成成本报表图片")
    parser.add_argument("--render-type", choices=["table", "chart"], default="chart")
    parser.add_argument("--report-id", required=True)
    parser.add_argument("--caller", default="debug")
    parser.add_argument("--user-id", default=None)
    args = parser.parse_args()

    try:
        save_path = generate(args.render_type, args.report_id, args.caller, args.user_id)
    except httpx.HTTPError as e:
        print(
            f"拉取成本数据失败：{e}\n"
            f"请确认模拟数据服务已启动且 THIRD_APP_BASE_URL（当前为 {THIRD_APP_BASE_URL!r}）配置正确。",
            file=sys.stderr,
        )
        print(f"ERROR:拉取成本数据失败：{e}")
        return 1
    except ReportRenderError as e:
        print(str(e), file=sys.stderr)
        print(f"ERROR:{e}")
        return 1
    except Exception as e:  # noqa: BLE001 - 兜底捕获，见文件顶部输出契约说明
        traceback.print_exc()
        print(f"ERROR:脚本内部错误：{e}")
        return 1

    print(f"RESULT_PATH:{save_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
