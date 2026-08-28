#!/usr/bin/env python3
"""支付宝匹配数据表生成 —— 沙箱容器内跑的 CLI 版本，供 deepagents 内置的 execute 工具调用。

跟原来 src/agent/tools/alipay_report_tools.py（@tool 版本，已废弃/待删）流程完全一致：分页
拉某一天（--date 不传则默认当天）全部支付宝匹配流水 -> 按 数据类别 -> 明细分类 -> 金额区间
三级动态分组聚合 -> 用 openpyxl 现场建表 -> 用 LibreOffice 把这张 sheet 整体截图成 PNG。
区别只是参数从 Python 对象（ContextSchema）换成命令行字符串参数，且本文件是自包含的——不
import 项目源码（容器里也没有那份代码，只有 skills/ 这一份 bind mount），LibreOffice 截图
原语（ReportRenderError/_soffice_convert 等）、命名/快照目录约定、分组聚合/建表逻辑都在
本文件内重新实现一份。

跟 cost-report/scripts/generate.py 的关键差异：
- 没有固定模板——每次调用用 openpyxl 从空白 workbook 现场建表，行数随当天数据量变化，
  也不需要 render_type（只有一种报表形态）。
- 类别/明细分类的取值集合与顺序完全由本次拉到的数据决定（首次出现顺序），本文件不允许
  出现任何写死的分类枚举——第三方接口随时可能返回开发者预先不知道的新类别。
- 不依赖 LibreOffice 重算公式——这份报表本身不含公式，是纯数值结果，openpyxl 原生写入即为
  最终结果；LibreOffice 在这里只用来把生成的 sheet 转 PDF 截图。

输出契约（调用方靠这个判断成功/失败，不看 exit_code 之外的东西，跟 cost-report 完全一致）：
- 成功：最后一行 stdout 打印 `RESULT_PATH:<容器内绝对路径>`，exit 0。
- 已知失败（拉数据失败、没有任何记录、LibreOffice 不可用/转换失败）：最后一行 stdout 打印
  `ERROR:<中文说明>`，exit 1。LibreOffice 自己往 stderr 打的噪声不会污染这行。
- 未预期的异常：完整 traceback 打到 stderr（供开发排查），stdout 最后一行仍是简化的
  `ERROR:脚本内部错误：<e>`，exit 1。
"""
from __future__ import annotations

import argparse
import io
import json
import os
import random
import re
import shutil
import subprocess
import sys
import tempfile
import traceback
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Literal

import httpx
import openpyxl
import pymupdf  # 把 LibreOffice 转出来的 PDF 页面渲成图片
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from PIL import Image, ImageDraw, ImageFont

SCRIPT_DIR = Path(__file__).resolve().parent
SKILL_DIR = SCRIPT_DIR.parent
# 容器里固定是 /workspace/output|snapshots；env 覆盖只用于宿主机本地验证（对齐
# cost-report/scripts/generate.py 的做法），不作为对外契约。
OUTPUT_DIR = Path(os.getenv("ALIPAY_REPORT_OUTPUT_DIR", "/workspace/output"))
SNAPSHOT_ROOT = Path(os.getenv("ALIPAY_REPORT_SNAPSHOT_ROOT", "/workspace/snapshots"))
FONT_PATH = SKILL_DIR / "assets" / "NotoSansSC-Regular.ttf"

SOFFICE_BIN = os.getenv("SOFFICE_BIN", "soffice")
THIRD_APP_BASE_URL = os.getenv("THIRD_APP_BASE_URL", "http://host.docker.internal:8800")

_PAGE_SIZE = 500  # 接口 page_size 上限（server.py Query(..., le=500)）
_MAX_PAGES = 20  # 约 2000 条 / 每页 500 条 = 4 页；20 是防止 total 异常返回导致死循环的保护上限

COLUMNS = [
    "序号", "数据类别", "明细分类", "金额区间",
    "下单笔数", "成功笔数", "下单金额", "成功金额", "成功率", "备注",
]
N_COLS = len(COLUMNS)

TITLE_ROW = 1
HEADER_ROW = 2
DATA_FIRST_ROW = 3

SUCCESS_STATUS = "成功"

# (标签, 上界含/None=无上界)，区间连续——每一档下界就是上一档的上界（严格大于），不
# 单独存下界，避免像 (0,500)/(501,2000) 那样在 500~501 之间留出空隙导致 500.5 这类
# 小数金额无法命中任何一档。顺序固定，不受当次数据影响。
AMOUNT_BANDS: list[tuple[str, float | None]] = [
    ("0-500", 500),
    ("501-2000", 2000),
    ("2001-10000", 10000),
    ("10001-30000", 30000),
    ("30001以上", None),
]

_TITLE_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
_TITLE_FONT = Font(color="FFFFFF", bold=True, size=14)
_HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_HEADER_FONT = Font(bold=True)
_SUBTOTAL_FILL = PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid")
_TOTAL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

_AMOUNT_FORMAT = "#,##0.00"
_PERCENT_FORMAT = "0.0%"

# 固定手工列宽——不用 excel_tools._autosize：第 1 行是跨 10 列的合并长标题，喂给
# 按列独立测宽的 autosize 会把 A 列撑得离谱，跟 A 列实际内容（1~2 位数的序号）无关。
_COLUMN_WIDTHS = {1: 6, 2: 18, 3: 18, 4: 14, 5: 10, 6: 10, 7: 16, 8: 16, 9: 10, 10: 22}


class ReportRenderError(Exception):
    """渲染/重算阶段的内部错误，转成 ERROR: 输出，不让异常裸露到顶层。"""


# ---------------------------------------------------------------------------
# 命名/快照目录约定（_naming.py + _report_snapshot.py 的字符串参数版本，跟
# cost-report/scripts/generate.py 完全一致）
# ---------------------------------------------------------------------------

_UNSAFE_CHARS = re.compile(r'[\\/:*?"<>|\s]+')


def sanitize(text: str, max_len: int = 50) -> str:
    safe = _UNSAFE_CHARS.sub("_", text).strip("_")
    return safe[:max_len] or "untitled"


def sanitize_user_id(user_id: str) -> str:
    return sanitize(str(user_id).split("@")[0])


def build_stem(title_stem: str, caller: str, user_id: str | None) -> str:
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    random_suffix = f"{random.randint(0, 999999):06d}"
    safe_title = sanitize(title_stem)
    if caller == "whatsapp" and user_id:
        return f"{safe_title}_{timestamp}_{sanitize_user_id(user_id)}_{random_suffix}"
    return f"DEBUG_{safe_title}_{timestamp}_{random_suffix}"


def _snapshot_user_folder(caller: str, user_id: str | None) -> str:
    if user_id:
        return sanitize_user_id(user_id)
    return caller


def resolve_snapshot_dir(report_id: str, caller: str, user_id: str | None) -> Path:
    snapshot_dir = SNAPSHOT_ROOT / _snapshot_user_folder(caller, user_id) / report_id
    snapshot_dir.mkdir(parents=True, exist_ok=True)
    return snapshot_dir


def _resolve_output_dir(caller: str, user_id: str | None) -> Path:
    if caller == "whatsapp" and user_id:
        d = OUTPUT_DIR / sanitize_user_id(user_id)
        d.mkdir(parents=True, exist_ok=True)
        return d
    return OUTPUT_DIR


def _resolve_image_save_path(caller: str, user_id: str | None) -> Path:
    user_output_dir = _resolve_output_dir(caller, user_id)
    stem = build_stem("支付宝匹配数据表", caller, user_id)
    candidate = user_output_dir / f"{stem}.png"
    n = 1
    while candidate.exists():
        candidate = user_output_dir / f"{stem}({n}).png"
        n += 1
    return candidate


# ---------------------------------------------------------------------------
# LibreOffice/pymupdf 截图原语（_report_screenshot.py 原样搬，跟
# cost-report/scripts/generate.py 完全一致）
# ---------------------------------------------------------------------------


def _soffice_convert(input_path: Path, outdir: Path, profile_dir: Path, fmt: str) -> Path:
    """outdir/profile_dir 必须每次调用用全新目录（调用方靠 tempfile.TemporaryDirectory 保证），
    避免并发调用抢同一份 LibreOffice 用户配置锁——这条不变量不能在迁移时弄丢。
    """
    try:
        result = subprocess.run(
            [
                SOFFICE_BIN,
                "--headless",
                "--calc",
                "--convert-to",
                fmt,
                "--outdir",
                str(outdir),
                str(input_path),
                f"-env:UserInstallation=file://{profile_dir}",
            ],
            capture_output=True,
            text=True,
            timeout=60,
        )
    except FileNotFoundError as e:
        raise ReportRenderError(
            "未找到 soffice 命令，请确认容器已安装 LibreOffice。"
        ) from e
    except subprocess.TimeoutExpired as e:
        raise ReportRenderError("LibreOffice 转换超时。") from e

    if result.returncode != 0:
        raise ReportRenderError(f"LibreOffice 转换失败：{result.stderr or result.stdout}")

    converted = outdir / f"{input_path.stem}.{fmt}"
    if not converted.exists():
        raise ReportRenderError("LibreOffice 没有生成转换后的文件。")
    return converted


def convert_to_pdf(input_path: Path, outdir: Path, profile_dir: Path) -> Path:
    return _soffice_convert(input_path, outdir, profile_dir, "pdf")


def render_pdf_page_with_footer(pdf_path: Path, page_index: int, out_path: Path, footer_text: str) -> None:
    doc = pymupdf.open(pdf_path)
    pix = doc[page_index].get_pixmap(dpi=200)
    doc.close()
    section_img = Image.open(io.BytesIO(pix.tobytes("png"))).convert("RGB")

    footer_height = 28
    canvas = Image.new("RGB", (section_img.width, section_img.height + footer_height), "white")
    canvas.paste(section_img, (0, 0))
    draw = ImageDraw.Draw(canvas)
    font = ImageFont.truetype(str(FONT_PATH), 16)
    text_width = draw.textlength(footer_text, font=font)
    draw.text(
        ((section_img.width - text_width) / 2, section_img.height + 4),
        footer_text,
        font=font,
        fill="#666666",
    )
    canvas.save(out_path)


# ---------------------------------------------------------------------------
# 数据拉取（alipay_report_tools.py 原样搬）
# ---------------------------------------------------------------------------


def _fetch_all_alipay_records(
    client: httpx.Client, target_date: date | None
) -> tuple[list[dict], str, int]:
    """返回 (全部记录, 这批数据对应的自然日, 接口 total)。`date`/`total` 都是接口在每一页
    响应里都会带的顶层字段，跟分页无关，取第一页返回的即可。

    `target_date` 为 None 时不传 date 查询参数，接口自己默认当天；传了就原样转 ISO
    字符串传给接口的 `date` 参数（接口按这个日期返回那一天的整日数据）。
    """
    items: list[dict] = []
    report_date: str | None = None
    total = None
    page = 1
    params: dict[str, int | str] = {"page_size": _PAGE_SIZE}
    if target_date is not None:
        params["date"] = target_date.isoformat()
    while (total is None or len(items) < total) and page <= _MAX_PAGES:
        resp = client.get(
            f"{THIRD_APP_BASE_URL}/api/alipay/matching-records",
            params={**params, "page": page},
        )
        resp.raise_for_status()
        data = resp.json()
        total = data["total"]
        if report_date is None:
            report_date = data["date"]
        items.extend(data["items"])
        if not data["items"]:
            break
        page += 1
    return items, report_date, total


# ---------------------------------------------------------------------------
# 分组聚合 + 建表 + 截图（_alipay_report_render.py 原样搬。类别/明细分类的取值
# 集合和显示顺序 100% 由传入的 records 决定，不允许出现任何硬编码的分类列表）
# ---------------------------------------------------------------------------


@dataclass
class Agg:
    count: int
    success_count: int
    amount: float
    success_amount: float

    @property
    def rate(self) -> float | None:
        return self.success_count / self.count if self.count else None


@dataclass
class DataRow:
    row_kind: Literal["band", "subtotal", "total"]
    data_category: str
    # band/subtotal 行：该行所属明细分类的原文本；total 行：f"{data_category}合计"
    detail_category_label: str
    band_label: str  # "0-500" 等区间标签 / "小计" / "合计"
    count: int
    success_count: int
    amount: float
    success_amount: float
    rate: float | None
    remark: str | None = None


@dataclass
class MergeSpan:
    start: int  # plan.rows 下标（含）
    end: int  # plan.rows 下标（含）
    value: str


@dataclass
class SheetPlan:
    rows: list[DataRow] = field(default_factory=list)
    data_category_spans: list[MergeSpan] = field(default_factory=list)
    detail_category_spans: list[MergeSpan] = field(default_factory=list)


def band_label(order_amount: float) -> str:
    for label, high in AMOUNT_BANDS:
        if high is None or order_amount <= high:
            return label
    return AMOUNT_BANDS[-1][0]


def _aggregate(records: list[dict]) -> Agg:
    success = [r for r in records if r["status"] == SUCCESS_STATUS]
    return Agg(
        count=len(records),
        success_count=len(success),
        amount=sum(r["order_amount"] for r in records),
        success_amount=sum(r["order_amount"] for r in success),
    )


def group_records(records: list[dict]) -> dict[str, dict[str, list[dict]]]:
    """按 data_category -> detail_category 两级分组。

    普通 dict 的插入顺序即首次出现顺序，天然实现"类别顺序完全由数据决定"，不需要
    任何排序或预先配置的分类清单。
    """
    grouped: dict[str, dict[str, list[dict]]] = {}
    for rec in records:
        grouped.setdefault(rec["data_category"], {}).setdefault(rec["detail_category"], []).append(rec)
    return grouped


def build_sheet_plan(records: list[dict]) -> SheetPlan:
    grouped = group_records(records)
    plan = SheetPlan()

    for data_category, detail_map in grouped.items():
        dc_start = len(plan.rows)
        dc_all_records: list[dict] = []

        for detail_category, detail_records in detail_map.items():
            dc_all_records.extend(detail_records)
            detail_start = len(plan.rows)

            buckets: dict[str, list[dict]] = {label: [] for label, _ in AMOUNT_BANDS}
            for r in detail_records:
                buckets[band_label(r["order_amount"])].append(r)

            for label, _ in AMOUNT_BANDS:
                bucket = buckets[label]
                if not bucket:
                    continue  # 该区间当前没有记录：跳过，不出零值行
                agg = _aggregate(bucket)
                plan.rows.append(
                    DataRow(
                        row_kind="band",
                        data_category=data_category,
                        detail_category_label=detail_category,
                        band_label=label,
                        count=agg.count,
                        success_count=agg.success_count,
                        amount=agg.amount,
                        success_amount=agg.success_amount,
                        rate=agg.rate,
                    )
                )

            sub_agg = _aggregate(detail_records)  # 全量汇总，数值上等于可见区间行之和
            plan.rows.append(
                DataRow(
                    row_kind="subtotal",
                    data_category=data_category,
                    detail_category_label=detail_category,
                    band_label="小计",
                    count=sub_agg.count,
                    success_count=sub_agg.success_count,
                    amount=sub_agg.amount,
                    success_amount=sub_agg.success_amount,
                    rate=sub_agg.rate,
                )
            )

            detail_end = len(plan.rows) - 1
            plan.detail_category_spans.append(MergeSpan(detail_start, detail_end, detail_category))

        total_agg = _aggregate(dc_all_records)  # 该数据类别下所有明细分类简单相加
        plan.rows.append(
            DataRow(
                row_kind="total",
                data_category=data_category,
                detail_category_label=f"{data_category}合计",
                band_label="合计",
                count=total_agg.count,
                success_count=total_agg.success_count,
                amount=total_agg.amount,
                success_amount=total_agg.success_amount,
                rate=total_agg.rate,
            )
        )

        dc_end = len(plan.rows) - 1  # 含合计行本身，合计行视觉上仍属于该数据类别
        plan.data_category_spans.append(MergeSpan(dc_start, dc_end, data_category))

    return plan


def derive_title_text(report_date: str) -> str:
    """标题里的日期直接取接口返回的 date 字段（ISO 格式，代表这批数据是哪一天的），
    不再从每条记录的 occurred_at 反推——接口现在已经明确给出这批数据对应的自然日。"""
    d = date.fromisoformat(report_date)
    return f"支付宝匹配数据表（{d.month}月{d.day}日）整日数据"


def render_workbook(plan: SheetPlan, title_text: str, out_path: Path) -> None:
    wb = Workbook()
    ws = wb.worksheets[0]  # 用 worksheets[0] 而不是 wb.active——后者的类型标注是
    # Worksheet | None，会让下面每一处 ws.xxx 调用都被类型检查器误报；新建的
    # Workbook() 必定有且只有这一张默认 sheet，worksheets[0] 的类型标注是非 None 的。
    ws.title = "支付宝匹配数据表"

    ws.cell(row=TITLE_ROW, column=1, value=title_text)
    ws.merge_cells(start_row=TITLE_ROW, start_column=1, end_row=TITLE_ROW, end_column=N_COLS)
    title_cell = ws.cell(row=TITLE_ROW, column=1)
    title_cell.fill = _TITLE_FILL
    title_cell.font = _TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[TITLE_ROW].height = 26

    for col_idx, name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx, value=name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, row in enumerate(plan.rows):
        r = DATA_FIRST_ROW + i
        ws.cell(row=r, column=1, value=i + 1)  # 序号：全表连续计数，不合并
        ws.cell(row=r, column=4, value=row.band_label)
        ws.cell(row=r, column=5, value=row.count)
        ws.cell(row=r, column=6, value=row.success_count)
        c7 = ws.cell(row=r, column=7, value=row.amount)
        c7.number_format = _AMOUNT_FORMAT
        c8 = ws.cell(row=r, column=8, value=row.success_amount)
        c8.number_format = _AMOUNT_FORMAT
        c9 = ws.cell(row=r, column=9, value=row.rate)
        c9.number_format = _PERCENT_FORMAT
        ws.cell(row=r, column=10, value=row.remark)

        if row.row_kind == "total":
            ws.cell(row=r, column=3, value=row.detail_category_label)  # 合计行独立写，不参与合并

        fill = _SUBTOTAL_FILL if row.row_kind == "subtotal" else _TOTAL_FILL if row.row_kind == "total" else None
        if fill:
            for c in range(1, N_COLS + 1):
                ws.cell(row=r, column=c).fill = fill

    for span in plan.data_category_spans:
        top, bottom = DATA_FIRST_ROW + span.start, DATA_FIRST_ROW + span.end
        ws.cell(row=top, column=2, value=span.value)
        if bottom > top:
            ws.merge_cells(start_row=top, start_column=2, end_row=bottom, end_column=2)
        ws.cell(row=top, column=2).alignment = Alignment(horizontal="center", vertical="center")

    for span in plan.detail_category_spans:
        top, bottom = DATA_FIRST_ROW + span.start, DATA_FIRST_ROW + span.end
        ws.cell(row=top, column=3, value=span.value)
        if bottom > top:
            ws.merge_cells(start_row=top, start_column=3, end_row=bottom, end_column=3)
        ws.cell(row=top, column=3).alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, width in _COLUMN_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = f"A{DATA_FIRST_ROW}"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def render_screenshot(xlsx_path: Path, out_path: Path, footer_text: str, tmp_dir: Path) -> None:
    """把 render_workbook 生成的单一 sheet 原样截图成一张 PNG，贴上 footer_text 脚注。

    只有一张 sheet、没有公式、没有跨 sheet 引用——不需要像成本报表那样先拆 sheet/摊平
    公式，直接在整张已用区域上设 print_area + 强制单页缩放即可转 PDF 截图。
    """
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.worksheets[0]
    ws.print_area = f"A1:{get_column_letter(N_COLS)}{ws.max_row}"
    ws.print_options.gridLines = False
    ws.print_options.headings = False
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    ws.page_margins = PageMargins(left=0.2, right=0.2, top=0.2, bottom=0.2, header=0, footer=0)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    src_path = tmp_dir / "alipay_screenshot_src.xlsx"
    wb.save(src_path)

    pdf_outdir = tmp_dir / "pdf_out"
    pdf_profile = tmp_dir / "pdf_profile"
    pdf_outdir.mkdir(exist_ok=True)
    pdf_profile.mkdir(exist_ok=True)
    pdf_path = convert_to_pdf(src_path, pdf_outdir, pdf_profile)

    render_pdf_page_with_footer(pdf_path, 0, out_path, footer_text)


# ---------------------------------------------------------------------------
# 主流程
# ---------------------------------------------------------------------------


def _parse_target_date(date_str: str | None) -> date | None:
    """把 CLI 传入的 --date 转成 date 对象；None 表示不限定，接口默认当天。

    调用方（agent）已经负责把"昨天""3号"这类口语化表达换算成 ISO 日期字符串，这里
    只做格式校验和"不能是未来"的兜底校验——避免脚本本身对着一个不存在的未来日期
    生成一份看起来煞有介事的假报表。
    """
    if date_str is None:
        return None
    try:
        parsed = date.fromisoformat(date_str)
    except ValueError as e:
        raise ReportRenderError(f"日期格式不对：{date_str!r}，需要 YYYY-MM-DD 格式。") from e
    if parsed > date.today():
        raise ReportRenderError(f"不能查询未来的日期：{parsed.isoformat()}。")
    return parsed


def generate(report_id: str, caller: str, user_id: str | None, date_str: str | None) -> Path:
    target_date = _parse_target_date(date_str)
    snapshot_dir = resolve_snapshot_dir(report_id, caller, user_id)

    with httpx.Client(timeout=30.0) as client:
        records, report_date, total = _fetch_all_alipay_records(client, target_date)

    if not records:
        raise ReportRenderError("第三方数据服务当前没有返回任何支付宝匹配记录，无法生成报表。")

    data_json_path = snapshot_dir / "data.json"
    data_json_path.write_text(
        json.dumps({"date": report_date, "total": total, "items": records}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    plan = build_sheet_plan(records)
    title_text = derive_title_text(report_date)
    report_xlsx_path = snapshot_dir / "report.xlsx"
    render_workbook(plan, title_text, report_xlsx_path)

    footer_text = f"报表ID: {report_id}"
    save_path = _resolve_image_save_path(caller, user_id)
    with tempfile.TemporaryDirectory(prefix="alipay_report_") as tmp_root:
        render_screenshot(report_xlsx_path, save_path, footer_text, Path(tmp_root))

    # 最终发给用户的图片也备份一份进快照文件夹，跟 report.xlsx/data.json 放在一起。
    shutil.copy2(save_path, snapshot_dir / save_path.name)

    return save_path


def main() -> int:
    parser = argparse.ArgumentParser(description="生成支付宝匹配数据表报表图片")
    parser.add_argument("--report-id", required=True)
    parser.add_argument("--caller", default="debug")
    parser.add_argument("--user-id", default=None)
    parser.add_argument("--date", default=None, help="YYYY-MM-DD，不传默认当天")
    args = parser.parse_args()

    try:
        save_path = generate(args.report_id, args.caller, args.user_id, args.date)
    except httpx.HTTPError as e:
        print(
            f"拉取支付宝匹配数据失败：{e}\n"
            f"请确认模拟数据服务已启动且 THIRD_APP_BASE_URL（当前为 {THIRD_APP_BASE_URL!r}）配置正确。",
            file=sys.stderr,
        )
        print(f"ERROR:拉取支付宝匹配数据失败：{e}")
        return 1
    except ReportRenderError as e:
        print(str(e), file=sys.stderr)
        print(f"ERROR:{e}")
        return 1
    except Exception as e:  # noqa: BLE001 - 兜底捕获，见文件顶部输出契约说明
        traceback.print_exc()
        print(f"ERROR:脚本内部错误：{e}")
        return 1

    print(f"RESULT_PATH:{save_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
