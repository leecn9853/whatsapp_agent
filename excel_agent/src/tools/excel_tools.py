"""Excel 相关工具：列出/查看 input 与 output 目录下的表格，聚合数据、生成图表 sheet。

统一直接操作项目根目录下的 input/、output/（原理和 save_file.py 一致，绕开
CompositeBackend/FilesystemBackend——xlsx 是二进制文件，本来也不适合走文本型的
read_file/edit_file）。

处理链路：文件第一次被处理时（源文件在 input/），结果另存为 output/ 下一个新文件名；
之后对同一个输出文件的后续操作（比如先聚合再画图）会识别出它已经在 output/ 里，直接
原地更新，不会越叠越多份文件。
"""
from __future__ import annotations

from pathlib import Path
from typing import Literal

import openpyxl
import pandas as pd
from langchain.tools import ToolRuntime
from langchain_core.tools import tool
from openpyxl.chart import BarChart, LineChart, PieChart, Reference, ScatterChart
from openpyxl.chart.series_factory import SeriesFactory as Series
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.context import ContextSchema
from src.tools._naming import build_stem, sanitize_user_id

_PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
INPUT_DIR = _PROJECT_ROOT / "input"
OUTPUT_DIR = _PROJECT_ROOT / "output"
INPUT_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# webhook.py 用这个集合判断哪些工具调用产生的文件需要自动发给用户（返回值是文件路径）
OUTPUT_FILE_TOOL_NAMES = {"aggregate_excel_sheet", "create_chart_sheet"}

_HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)

_CHART_CLASSES = {
    "bar": BarChart,
    "line": LineChart,
    "pie": PieChart,
    "scatter": ScatterChart,
}


class ExcelToolError(Exception):
    """工具内部错误，被外层捕获后转成给 agent 看的错误文本，而不是抛异常中断整轮对话。"""


def _style_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autosize(ws) -> None:
    for col in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0 for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(10, length + 2)


def _user_dir(base: Path, ctx: ContextSchema | None) -> Path:
    """WhatsApp 调用按 user_id 隔离到子目录（自动创建），避免不同用户互相看到/
    处理到对方上传的文件；调试（直接跑 src/main.py）时返回 base 本身，行为不变，
    仍然使用 input/gen_reports.py 生成的根目录样例文件。
    """
    if ctx and ctx.caller == "whatsapp" and ctx.user_id:
        d = base / sanitize_user_id(ctx.user_id)
        d.mkdir(parents=True, exist_ok=True)
        return d
    return base


def _resolve_load_path(filename: str, ctx: ContextSchema | None) -> Path:
    """先找 output/（正在处理中的副本），再找 input/（原始样例/用户上传的文件）。"""
    name = Path(filename).name  # 只取文件名，防止路径穿越到其他目录
    for base in (_user_dir(OUTPUT_DIR, ctx), _user_dir(INPUT_DIR, ctx)):
        candidate = base / name
        if candidate.is_file():
            return candidate
    raise ExcelToolError(f"没有找到文件 {filename!r}，请先用 list_excel_files 确认文件名。")


def _resolve_save_path(load_path: Path, filename: str, output_filename: str | None, ctx: ContextSchema | None) -> Path:
    """已经在 output/ 里的直接原地覆盖；否则从 input/ 复制出一个新文件名。"""
    user_output_dir = _user_dir(OUTPUT_DIR, ctx)
    if load_path.parent == user_output_dir:
        return load_path

    if output_filename:
        name = Path(output_filename).name
        stem, suffix = Path(name).stem, (Path(name).suffix or ".xlsx")
    else:
        stem = build_stem(Path(filename).stem, ctx)
        suffix = Path(filename).suffix or ".xlsx"

    candidate = user_output_dir / f"{stem}{suffix}"
    n = 1
    while candidate.exists():
        candidate = user_output_dir / f"{stem}({n}){suffix}"
        n += 1
    return candidate


def save_uploaded_file(user_id: str, filename: str, content: bytes) -> Path:
    """把 WhatsApp 用户上传的文件存进该用户专属的 input/ 子目录。

    供 src/webhook/whatsapp.py 收到 document 消息时直接调用，不作为 @tool 暴露给
    LLM（LLM 不需要、也不应该自己决定往磁盘写用户上传的原始文件）。
    """
    target_dir = _user_dir(INPUT_DIR, ContextSchema(caller="whatsapp", user_id=user_id))
    name = Path(filename).name  # 防路径穿越
    stem, suffix = Path(name).stem, Path(name).suffix
    candidate = target_dir / name
    n = 1
    while candidate.exists():
        candidate = target_dir / f"{stem}({n}){suffix}"
        n += 1
    candidate.write_bytes(content)
    return candidate


def _list_sheets(path: Path) -> list[str]:
    wb = openpyxl.load_workbook(path, read_only=True)
    try:
        return wb.sheetnames
    finally:
        wb.close()


@tool
def list_excel_files(runtime: ToolRuntime) -> str:
    """列出 input/ 和 output/ 目录下所有可处理的 Excel 文件（.xlsx）及其 sheet 名称。

    动手处理某个文件前，如果不确定确切文件名/sheet 名，先调用这个工具确认。
    """
    lines = []
    for label, base in (
        ("input", _user_dir(INPUT_DIR, runtime.context)),
        ("output", _user_dir(OUTPUT_DIR, runtime.context)),
    ):
        files = sorted(base.glob("*.xlsx"))
        if not files:
            lines.append(f"[{label}] （空）")
            continue
        lines.append(f"[{label}]")
        for f in files:
            try:
                lines.append(f"  - {f.name}  sheets={_list_sheets(f)}")
            except Exception as e:
                lines.append(f"  - {f.name}  (读取失败: {e})")
    return "\n".join(lines)


@tool
def inspect_excel(filename: str, runtime: ToolRuntime, sheet_name: str | None = None) -> str:
    """查看某个 Excel 文件的结构：每个 sheet 的尺寸、表头、前几行数据预览。

    filename 只写文件名（不带目录），会自动先在 output/ 里找正在处理的副本，
    找不到再去 input/ 找原始文件。sheet_name 留空则列出所有 sheet 的概览；
    指定某个 sheet 名则展开显示该 sheet 更完整的预览（最多 10 行数据）。
    做任何聚合/画图之前，建议先用这个工具看清楚表头和数据形态。
    """
    try:
        path = _resolve_load_path(filename, runtime.context)
    except ExcelToolError as e:
        return str(e)

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name and sheet_name not in wb.sheetnames:
            return f"sheet {sheet_name!r} 不存在，该文件的 sheet 有：{wb.sheetnames}"
        target_sheets = [sheet_name] if sheet_name else wb.sheetnames

        blocks = []
        for name in target_sheets:
            ws = wb[name]
            n_rows, n_cols = ws.max_row, ws.max_column
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                blocks.append(f"## {name}\n(空 sheet)")
                continue
            header, data_rows = rows[0], rows[1:11]
            table = ["| " + " | ".join(str(h) for h in header) + " |",
                     "| " + " | ".join("---" for _ in header) + " |"]
            for r in data_rows:
                table.append("| " + " | ".join("" if v is None else str(v) for v in r) + " |")
            more = f"\n（仅展示前 {len(data_rows)} 行数据，共 {n_rows - 1} 行）" if n_rows - 1 > len(data_rows) else ""
            blocks.append(f"## {name}\n共 {n_rows} 行 x {n_cols} 列（含表头）\n\n" + "\n".join(table) + more)
        return f"文件：{path.name}\n\n" + "\n\n".join(blocks)
    finally:
        wb.close()


@tool(response_format="content_and_artifact")
def aggregate_excel_sheet(
    filename: str,
    sheet_name: str,
    group_by: list[str],
    aggregations: dict[str, str],
    new_sheet_name: str,
    runtime: ToolRuntime,
    output_filename: str | None = None,
) -> tuple[str, str | None]:
    """按列分组聚合某个 sheet 的数据，把聚合结果写成同一个工作簿里的新 sheet。

    数据粒度很细（比如"月份 x 产品线 x 地区"逐行）时，直接对原始行画图会很难看，
    应该先用这个工具聚合出一张干净的汇总表，再用 create_chart_sheet 基于这张汇总表出图。

    - group_by：用于分组的列名列表（必须是表头里已有的列名）。
    - aggregations：{列名: "sum"|"mean"|"count"}，指定要聚合哪些数值列、用什么方式。
    - new_sheet_name：写入的新 sheet 名称，若已存在同名 sheet 会被覆盖。
    - output_filename：可选，指定另存的文件名；不指定则按处理链路自动决定（详见模块说明）。

    返回：实际写入的文件名，以及聚合结果的预览。
    """
    try:
        load_path = _resolve_load_path(filename, runtime.context)
    except ExcelToolError as e:
        return str(e), None

    try:
        df = pd.read_excel(load_path, sheet_name=sheet_name)
    except Exception as e:
        return f"读取 sheet {sheet_name!r} 失败：{e}", None

    missing = [c for c in [*group_by, *aggregations] if c not in df.columns]
    if missing:
        return f"列名不存在：{missing}；该 sheet 的列有：{list(df.columns)}", None

    try:
        grouped = df.groupby(group_by, as_index=False).agg(aggregations)
    except Exception as e:
        return f"聚合失败：{e}", None

    wb = openpyxl.load_workbook(load_path)
    if new_sheet_name in wb.sheetnames:
        del wb[new_sheet_name]
    ws = wb.create_sheet(new_sheet_name)
    ws.append(list(grouped.columns))
    for row in grouped.itertuples(index=False):
        ws.append(list(row))
    _style_header(ws)
    _autosize(ws)

    save_path = _resolve_save_path(load_path, filename, output_filename, runtime.context)
    wb.save(save_path)

    preview = grouped.head(10).to_string(index=False)
    content = (
        f"已生成汇总 sheet「{new_sheet_name}」，写入文件：{save_path.name}\n\n"
        f"聚合结果预览：\n{preview}"
    )
    return content, str(save_path)


@tool(response_format="content_and_artifact")
def create_chart_sheet(
    filename: str,
    sheet_name: str,
    chart_type: Literal["bar", "line", "pie", "scatter"],
    category_column: str,
    value_columns: list[str],
    title: str,
    runtime: ToolRuntime,
    new_sheet_name: str = "图表",
    output_filename: str | None = None,
) -> tuple[str, str | None]:
    """在 Excel 文件里新建一个 sheet，基于某个已有 sheet 的数据插入原生 Excel 图表。

    图表是真正的 Excel 图表对象（不是图片），用户在 Excel/WPS 打开后仍可编辑、跟数据联动。
    建议先用 inspect_excel 看清表头，数据粒度太细时先用 aggregate_excel_sheet 聚合出一张
    干净的汇总表，再基于那张表调用本工具，否则图表会因为类别/数据点太多而不可读。

    - chart_type：bar(柱状图，类别对比) / line(折线图，时间序列) / pie(饼图，占比，
      只能传一个 value_column，类别数建议 ≤6) / scatter(散点图，两个数值变量的相关性)。
    - category_column：作为图表横轴/类别的列名（必须是该 sheet 表头里已有的列名）。
    - value_columns：要画成数据系列的数值列名列表；pie 类型只能传恰好一个。
    - new_sheet_name：新建的图表 sheet 名称，若已存在同名 sheet 会被覆盖。

    返回：实际写入的文件名。
    """
    if chart_type == "pie" and len(value_columns) != 1:
        return "pie（饼图）只支持一个 value_column，Excel 饼图语义上不能画多个数值系列。", None

    try:
        load_path = _resolve_load_path(filename, runtime.context)
    except ExcelToolError as e:
        return str(e), None

    wb = openpyxl.load_workbook(load_path)
    if sheet_name not in wb.sheetnames:
        return f"sheet {sheet_name!r} 不存在，该文件的 sheet 有：{wb.sheetnames}", None
    data_ws = wb[sheet_name]

    header = [c.value for c in next(data_ws.iter_rows(min_row=1, max_row=1))]
    if category_column not in header:
        return f"列名 {category_column!r} 不存在，该 sheet 的列有：{header}", None
    missing_values = [c for c in value_columns if c not in header]
    if missing_values:
        return f"列名不存在：{missing_values}；该 sheet 的列有：{header}", None

    max_row = data_ws.max_row
    cat_col_idx = header.index(category_column) + 1
    value_col_idxs = [header.index(c) + 1 for c in value_columns]

    chart = _CHART_CLASSES[chart_type]()
    chart.title = title
    chart.style = 10

    if chart_type == "scatter":
        xvalues = Reference(data_ws, min_col=cat_col_idx, min_row=2, max_row=max_row)
        for idx in value_col_idxs:
            yvalues = Reference(data_ws, min_col=idx, min_row=1, max_row=max_row)
            chart.series.append(Series(yvalues, xvalues, title_from_data=True))
    else:
        cats = Reference(data_ws, min_col=cat_col_idx, min_row=2, max_row=max_row)
        for idx in value_col_idxs:
            data = Reference(data_ws, min_col=idx, min_row=1, max_row=max_row)
            chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

    if chart_type != "pie":
        chart.x_axis.title = category_column
        chart.y_axis.title = "、".join(value_columns)

    if chart_type == "pie" or len(value_columns) > 1:
        chart.legend.position = "b"
    else:
        chart.legend = None

    if new_sheet_name in wb.sheetnames:
        del wb[new_sheet_name]
    chart_ws = wb.create_sheet(new_sheet_name)
    chart_ws.add_chart(chart, "B2")

    save_path = _resolve_save_path(load_path, filename, output_filename, runtime.context)
    wb.save(save_path)
    content = f"已在「{new_sheet_name}」sheet 里插入 {chart_type} 图表，写入文件：{save_path.name}"
    return content, str(save_path)
